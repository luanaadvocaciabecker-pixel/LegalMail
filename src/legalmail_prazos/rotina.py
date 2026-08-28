"""Orquestração das partes 1 e 2 da rotina (seções 4, 6, 8, 9 e 10).

Este módulo assume que a interpretação jurídica de cada intimação (qual é
o ato exato, quantos dias tem o prazo, se comporta embargos etc. — seção 5,
itens 4, 5 e 11) já foi feita por quem lê o teor da intimação (pessoa ou
agente com acesso ao texto), usando :mod:`legalmail_prazos.prazos` para o
cálculo determinístico das datas. O que este módulo mecaniza é a parte
verificável sem interpretação de linguagem natural: decidir caso novo vs.
recorrente, escrever a planilha com segurança, criar a tarefa reaproveitando
o tipo do histórico quando possível, arquivar da Entrada para o Acervo, e
montar o relatório final no formato exigido pela seção 10.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from .classificacao import (
    contem_designacao_pericia,
    eh_intimacao_de_audiencia,
    extrair_data_e_horario,
    sugerir_area_por_tribunal,
    sugerir_evento_audiencia,
)
from .legalmail_client import (
    AudienciaLegalmail,
    ItemEntrada,
    LegalmailClient,
    NovaTarefaLegalmail,
    TipoTarefaHistorico,
)
from .planilha import (
    ABA_AUDIENCIA,
    ABA_PRAZOS,
    NovaAudiencia,
    NovaPericia,
    NovoCasoPrazos,
    adicionar_linha_audiencia,
    adicionar_linha_pericia,
    adicionar_linha_prazos,
    carregar_processos_conhecidos,
    conferir_celula,
    eh_caso_novo,
    fazer_backup,
    localizar_advogado_por_processo,
    normalizar_processo,
    processos_na_aba_audiencia,
    processos_na_aba_prazos,
    salvar_com_seguranca,
    salvar_processos_conhecidos,
    COL_PRAZOS_PROCESSO,
)
from .prazos import Prazo


def escolher_tipo_tarefa(
    historico: list[TipoTarefaHistorico], descricao_nova: str
) -> str | None:
    """Seção 6: tenta reaproveitar um Tipo já usado no histórico do processo.

    Heurística simples por sobreposição de palavras entre a descrição da
    nova tarefa e o resumo de tarefas anteriores. Não decide por conta
    própria quando não há correspondência razoável: nesses casos retorna
    ``None`` para que o Tipo seja confirmado manualmente antes de criar um
    Tipo novo no Legalmail (nunca criar "(novo)" sem necessidade real).
    """

    if not historico:
        return None

    def palavras(texto: str) -> set[str]:
        return {p for p in texto.upper().split() if len(p) > 3}

    alvo = palavras(descricao_nova)
    melhor: tuple[int, str] | None = None
    for item in historico:
        score = len(alvo & palavras(item.descricao_resumo))
        if score > 0 and (melhor is None or score > melhor[0]):
            melhor = (score, item.tipo)
    return melhor[1] if melhor else None


def triar_entrada(itens: list[ItemEntrada]) -> tuple[list[ItemEntrada], list[ItemEntrada]]:
    """Separa a Entrada em (candidatas a audiência, demais itens) — seção 9.

    A API não tem endpoint de audiências; a única forma de identificá-las é
    classificar as próprias intimações da Entrada
    (``classificacao.eh_intimacao_de_audiencia``). Os "demais itens" seguem
    o fluxo normal da Parte 1 (:func:`processar_parte1`).
    """

    audiencias: list[ItemEntrada] = []
    demais: list[ItemEntrada] = []
    for item in itens:
        if eh_intimacao_de_audiencia(tipo=item.tipo, teor=item.conteudo_intimacao):
            audiencias.append(item)
        else:
            demais.append(item)
    return audiencias, demais


def sugerir_pericia(item: ItemEntrada) -> NovaPericia | None:
    """Seção 4, item 4: sugere uma entrada de PERÍCIA quando o teor da
    intimação menciona uma designação (vale mesmo para casos recorrentes).

    Retorna ``None`` quando o teor não menciona perícia. Data e horário só
    vêm preenchidos quando extraídos do teor com uma correspondência clara
    (ver :func:`legalmail_prazos.classificacao.extrair_data_e_horario`);
    caso contrário ficam em branco na planilha, para preenchimento manual.
    """

    if not contem_designacao_pericia(item.conteudo_intimacao):
        return None
    extracao = extrair_data_e_horario(item.conteudo_intimacao)
    return NovaPericia(
        cliente=item.cliente_x_parte,
        numero_processo=item.numero_processo,
        pericia="A CONFIRMAR",
        data=extracao.data,
        horario=extracao.horario,
    )


@dataclass
class DecisaoItemEntrada:
    """Fatos já resolvidos sobre um item da Entrada, prontos para aplicar as regras."""

    item: ItemEntrada
    descricao_tarefa: str
    prazo: Prazo
    tipo_tarefa_sugerido: str | None = None
    """Se já souber o Tipo a reaproveitar do histórico, informe aqui;
    caso contrário deixe None para tentar :func:`escolher_tipo_tarefa`."""
    encarregado_override: str | None = None
    pericia: NovaPericia | None = None
    observacoes_prazos: str | None = None


@dataclass
class ResultadoCasoNovo:
    numero_processo: str
    cliente: str
    tribunal: str


@dataclass
class RelatorioConciliacao:
    casos_novos: list[ResultadoCasoNovo] = field(default_factory=list)
    tarefas_criadas: int = 0
    processos_encarregados: int = 0
    pericias_adicionadas: int = 0
    audiencias_adicionadas: int = 0
    limitacoes: list[str] = field(default_factory=list)
    tipos_pendentes_confirmacao: list[str] = field(default_factory=list)

    def texto(self) -> str:
        """Monta o relatório em parágrafos corridos, sem dois-pontos, bullets ou tabelas (seção 10)."""

        partes: list[str] = []
        if self.casos_novos:
            descricoes = "; ".join(
                f"{c.cliente} processo {c.numero_processo} ({c.tribunal})" for c in self.casos_novos
            )
            partes.append(
                f"Foram cadastrados e arquivados para o Acervo {len(self.casos_novos)} "
                f"casos novos, sendo eles {descricoes}."
            )
        else:
            partes.append("Não foram encontrados casos novos na Entrada nesta execução.")

        partes.append(
            f"Foram criadas {self.tarefas_criadas} tarefas no Legalmail com os prazos calculados."
        )

        if self.processos_encarregados:
            partes.append(
                f"Foram encarregados {self.processos_encarregados} processos ao advogado "
                "responsável correspondente."
            )

        if self.pericias_adicionadas:
            partes.append(
                f"Foram adicionadas {self.pericias_adicionadas} perícias designadas à aba PERÍCIA."
            )

        if self.audiencias_adicionadas:
            partes.append(
                f"Foram adicionadas {self.audiencias_adicionadas} audiências futuras ainda não "
                "cadastradas à aba AUDIÊNCIA."
            )

        if self.tipos_pendentes_confirmacao:
            partes.append(
                "Os seguintes itens tiveram o Tipo da tarefa criado sem correspondência clara no "
                "histórico do processo e precisam de confirmação manual, a saber "
                + "; ".join(self.tipos_pendentes_confirmacao)
                + "."
            )

        if self.limitacoes:
            partes.append(
                "A execução foi interrompida ou parcialmente limitada nos seguintes pontos, "
                + "; ".join(self.limitacoes)
                + "."
            )

        return " ".join(partes)


def processar_parte1(
    *,
    decisoes: list[DecisaoItemEntrada],
    client: LegalmailClient,
    caminho_planilha: Path,
    caminho_prazos_nums_json: Path,
    pasta_outputs: Path,
    mapa_abreviacao_para_nome_completo: dict[str, str] | None = None,
) -> RelatorioConciliacao:
    """Executa a Parte 1 (seções 4, 6, 7, 8) sobre uma lista de decisões já resolvidas."""

    relatorio = RelatorioConciliacao()
    linhas_criadas: list[tuple[str, int]] = []

    fazer_backup(caminho_planilha, pasta_outputs)
    wb = load_workbook(caminho_planilha, data_only=False)
    ws_prazos = wb[ABA_PRAZOS]

    processos_conhecidos = carregar_processos_conhecidos(caminho_prazos_nums_json)
    processos_conhecidos |= processos_na_aba_prazos(ws_prazos)

    for decisao in decisoes:
        item = decisao.item

        if not item.conteudo_acessivel:
            relatorio.limitacoes.append(
                f"o conteúdo do processo {item.numero_processo} está bloqueado (segredo de "
                "justiça sem acesso real) e não pôde ser processado"
            )
            continue

        numero_normalizado = normalizar_processo(item.numero_processo)
        novo = eh_caso_novo(numero_normalizado, processos_conhecidos)

        advogado = decisao.encarregado_override or localizar_advogado_por_processo(
            wb,
            item.numero_processo,
            mapa_abreviacao_para_nome_completo=mapa_abreviacao_para_nome_completo,
        )
        if advogado is None:
            relatorio.limitacoes.append(
                f"não foi possível identificar o advogado responsável pelo processo "
                f"{item.numero_processo} na aba ATIVOS ATUAL"
            )
            continue

        if novo:
            linha_criada = adicionar_linha_prazos(
                wb,
                NovoCasoPrazos(
                    tribunal=item.tribunal,
                    numero_processo=item.numero_processo,
                    cliente_x_parte=item.cliente_x_parte,
                    advogado=advogado,
                    segredo_justica=item.segredo_justica,
                    observacoes=decisao.observacoes_prazos,
                ),
            )
            processos_conhecidos.add(numero_normalizado)
            linhas_criadas.append((item.numero_processo, linha_criada))
            relatorio.casos_novos.append(
                ResultadoCasoNovo(
                    numero_processo=item.numero_processo,
                    cliente=item.cliente_x_parte,
                    tribunal=item.tribunal,
                )
            )

        if decisao.pericia is not None:
            adicionar_linha_pericia(wb, decisao.pericia)
            relatorio.pericias_adicionadas += 1

        try:
            historico = client.historico_tarefas_do_processo(item.id_legalmail)
        except NotImplementedError:
            historico = []

        tipo = decisao.tipo_tarefa_sugerido or escolher_tipo_tarefa(
            historico, decisao.descricao_tarefa
        )
        if tipo is None:
            relatorio.tipos_pendentes_confirmacao.append(
                f"{item.cliente_x_parte} processo {item.numero_processo}"
            )
            tipo = "(a confirmar)"

        try:
            client.criar_tarefa(
                NovaTarefaLegalmail(
                    id_legalmail_processo=item.id_legalmail,
                    tipo=tipo,
                    encarregado=advogado,
                    descricao=decisao.descricao_tarefa,
                    prazo_legal=decisao.prazo.data_final_legal,
                    prazo_interno=decisao.prazo.data_final_interna,
                )
            )
            relatorio.tarefas_criadas += 1
        except NotImplementedError as exc:
            relatorio.limitacoes.append(
                f"não foi possível criar a tarefa de {item.cliente_x_parte} processo "
                f"{item.numero_processo} ({exc}), lançar manualmente na interface do Legalmail"
            )

        # Encaminha ao responsável mesmo quando o backend não suporta criar tarefa
        # (ver seção 8 e a nota sobre encarregar_advogado no contrato LegalmailClient).
        try:
            id_usuario = client.localizar_usuario_por_nome(advogado)
        except NotImplementedError:
            id_usuario = None
        if id_usuario is None:
            relatorio.limitacoes.append(
                f"não foi possível localizar o usuário do Legalmail correspondente a {advogado} "
                f"para encarregar o processo {item.numero_processo}"
            )
        else:
            try:
                client.encarregar_advogado(item.id_legalmail, id_usuario)
                relatorio.processos_encarregados += 1
            except NotImplementedError as exc:
                relatorio.limitacoes.append(
                    f"não foi possível encarregar {advogado} pelo processo {item.numero_processo} "
                    f"({exc})"
                )

        client.arquivar_para_acervo(item.id_legalmail)

    salvar_com_seguranca(wb, caminho_planilha)
    salvar_processos_conhecidos(caminho_prazos_nums_json, processos_conhecidos)

    for numero_processo, linha_criada in linhas_criadas:
        ok = conferir_celula(
            caminho_planilha,
            ABA_PRAZOS,
            linha=linha_criada,
            coluna=COL_PRAZOS_PROCESSO,
            valor_esperado=numero_processo,
        )
        if not ok:
            relatorio.limitacoes.append(
                f"a gravação do processo {numero_processo} não pôde ser confirmada após "
                "reabrir o arquivo"
            )

    return relatorio


def processar_parte2(
    *,
    audiencias_legalmail: list[AudienciaLegalmail],
    caminho_planilha: Path,
    pasta_outputs: Path,
    relatorio: RelatorioConciliacao | None = None,
) -> RelatorioConciliacao:
    """Executa a Parte 2 (seção 9): audiências futuras ainda não cadastradas."""

    relatorio = relatorio or RelatorioConciliacao()

    fazer_backup(caminho_planilha, pasta_outputs)
    wb = load_workbook(caminho_planilha, data_only=False)
    ws_audiencia = wb[ABA_AUDIENCIA]

    ja_cadastradas = processos_na_aba_audiencia(ws_audiencia)

    for audiencia in audiencias_legalmail:
        numero_normalizado = normalizar_processo(audiencia.numero_processo)
        if numero_normalizado in ja_cadastradas:
            continue

        evento = audiencia.evento
        if audiencia.cancelada and "CANCELAD" not in evento.upper():
            evento = f"{evento} (CANCELADA)"

        adicionar_linha_audiencia(
            wb,
            NovaAudiencia(
                cliente=audiencia.cliente,
                numero_processo=audiencia.numero_processo,
                evento=evento,
                area=audiencia.area,
                data=audiencia.data,
                horario=audiencia.horario.strftime("%H:%M"),
            ),
        )
        ja_cadastradas.add(numero_normalizado)
        relatorio.audiencias_adicionadas += 1

    salvar_com_seguranca(wb, caminho_planilha)
    return relatorio


def escrever_audiencias_da_entrada(
    *,
    itens_audiencia: list[ItemEntrada],
    caminho_planilha: Path,
    pasta_outputs: Path,
    relatorio: RelatorioConciliacao | None = None,
) -> RelatorioConciliacao:
    """Escreve na aba AUDIÊNCIA as intimações da Entrada já classificadas como
    audiência por :func:`triar_entrada` (seção 9).

    Diferente de :func:`processar_parte2` (que espera audiências já
    conhecidas, com data e horário certos, vindas de ``listar_audiencias``),
    esta função extrai data e horário do teor da própria intimação —
    best-effort — e deixa em branco quando a extração não é confiável,
    sinalizando isso no relatório em vez de arriscar uma data errada.
    """

    relatorio = relatorio or RelatorioConciliacao()

    fazer_backup(caminho_planilha, pasta_outputs)
    wb = load_workbook(caminho_planilha, data_only=False)
    ws_audiencia = wb[ABA_AUDIENCIA]

    ja_cadastradas = processos_na_aba_audiencia(ws_audiencia)

    for item in itens_audiencia:
        numero_normalizado = normalizar_processo(item.numero_processo)
        if numero_normalizado in ja_cadastradas:
            continue

        extracao = extrair_data_e_horario(item.conteudo_intimacao)
        adicionar_linha_audiencia(
            wb,
            NovaAudiencia(
                cliente=item.cliente_x_parte,
                numero_processo=item.numero_processo,
                evento=sugerir_evento_audiencia(item.conteudo_intimacao),
                area=sugerir_area_por_tribunal(item.tribunal),
                data=extracao.data,
                horario=extracao.horario,
            ),
        )
        ja_cadastradas.add(numero_normalizado)
        relatorio.audiencias_adicionadas += 1
        if extracao.confirmar_manualmente:
            relatorio.limitacoes.append(
                f"a data/horário da audiência de {item.cliente_x_parte} processo "
                f"{item.numero_processo} não pôde ser extraído do teor com confiança "
                "e ficou em branco, confirmar manualmente"
            )

    salvar_com_seguranca(wb, caminho_planilha)
    return relatorio
