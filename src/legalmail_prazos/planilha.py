"""Leitura e escrita da planilha PRAZOS BECKER (seções 2 a 4 e 9).

Este módulo só sabe falar com a planilha .xlsx; ele não sabe nada sobre a
UI do Legalmail. Isso é intencional (seção 11 do documento de referência):
a lógica de negócio deve ser independente da camada usada para obter os
dados de entrada (API, exportação CSV, ou automação de navegador).
"""

from __future__ import annotations

import json
import os
import re
import shutil
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ABA_PRAZOS = "PRAZOS"
ABA_ATIVOS_ATUAL = "ATIVOS ATUAL"
ABA_PERICIA = "PERÍCIA"
ABA_AUDIENCIA = "AUDIÊNCIA"

# Linha do cabeçalho de dados na aba PRAZOS (as linhas acima são o resumo/legenda).
LINHA_CABECALHO_PRAZOS = 8
COL_PRAZOS_TRIBUNAL = 1  # A
COL_PRAZOS_PROCESSO = 2  # B
COL_PRAZOS_CLIENTE = 3  # C
COL_PRAZOS_PRAZO1 = 4  # D
COL_PRAZOS_PRAZO2 = 5  # E
COL_PRAZOS_INTERNO = 6  # F
COL_PRAZOS_ADVOGADO = 7  # G
COL_PRAZOS_STATUS = 8  # H
COL_PRAZOS_OBSERVACOES = 9  # I

COL_ATIVOS_PROCESSO = 1  # A
COL_ATIVOS_ADVOGADO_ATUAL = 11  # K


def normalizar_processo(numero: str) -> str:
    """Remove toda pontuação/espaço e mantém só os dígitos do número do processo.

    Trata também o caractere non-breaking space (``\xa0``) observado em
    exportações do Legalmail antes do número do processo.
    """

    return re.sub(r"\D", "", numero or "")


def carregar_processos_conhecidos(caminho_prazos_nums_json: Path) -> set[str]:
    """Carrega ``prazos_nums.json`` (lista de números já normalizados)."""

    if not caminho_prazos_nums_json.exists():
        return set()
    with caminho_prazos_nums_json.open("r", encoding="utf-8") as f:
        dados = json.load(f)
    return {normalizar_processo(n) for n in dados}


def salvar_processos_conhecidos(caminho: Path, processos: set[str]) -> None:
    with caminho.open("w", encoding="utf-8") as f:
        json.dump(sorted(processos), f, ensure_ascii=False, indent=2)


def processos_na_aba_prazos(ws: Worksheet) -> set[str]:
    """Números de processo (normalizados) já presentes na coluna B da aba PRAZOS."""

    encontrados: set[str] = set()
    for linha in range(LINHA_CABECALHO_PRAZOS + 1, ws.max_row + 1):
        valor = ws.cell(row=linha, column=COL_PRAZOS_PROCESSO).value
        if valor:
            encontrados.add(normalizar_processo(str(valor)))
    return encontrados


def eh_caso_novo(numero_processo: str, processos_conhecidos: set[str]) -> bool:
    """Seção 4: um caso é recorrente se o número normalizado já é conhecido."""

    return normalizar_processo(numero_processo) not in processos_conhecidos


def localizar_advogado_por_processo(
    wb: Workbook,
    numero_processo: str,
    *,
    mapa_abreviacao_para_nome_completo: dict[str, str] | None = None,
) -> str | None:
    """Cruza o processo com a aba ATIVOS ATUAL para achar o advogado responsável.

    A coluna K dessa aba costuma trazer só o primeiro nome/abreviação (ex.
    "Alana"), não o nome completo usado na coluna G da aba PRAZOS (ex.
    "ALANA PAIS LEMOS"). Se ``mapa_abreviacao_para_nome_completo`` for
    informado e contiver a abreviação encontrada, o nome completo é
    retornado; caso contrário devolve-se o valor bruto da planilha, para
    nunca inventar um nome que não veio da fonte de dados (seção 7).
    """

    if ABA_ATIVOS_ATUAL not in wb.sheetnames:
        return None
    ws = wb[ABA_ATIVOS_ATUAL]
    alvo = normalizar_processo(numero_processo)
    for linha in range(2, ws.max_row + 1):
        valor = ws.cell(row=linha, column=COL_ATIVOS_PROCESSO).value
        if valor and normalizar_processo(str(valor)) == alvo:
            bruto = ws.cell(row=linha, column=COL_ATIVOS_ADVOGADO_ATUAL).value
            if not bruto:
                return None
            bruto = str(bruto).strip()
            if mapa_abreviacao_para_nome_completo:
                return mapa_abreviacao_para_nome_completo.get(bruto, bruto)
            return bruto
    return None


def _proxima_linha_livre(ws: Worksheet, coluna_referencia: int) -> int:
    linha = ws.max_row
    while linha > 1 and ws.cell(row=linha, column=coluna_referencia).value in (None, ""):
        linha -= 1
    return linha + 1


@dataclass
class NovoCasoPrazos:
    tribunal: str
    numero_processo: str
    cliente_x_parte: str
    advogado: str
    segredo_justica: bool = False
    observacoes: str | None = None


def adicionar_linha_prazos(wb: Workbook, caso: NovoCasoPrazos) -> int:
    """Seção 4 (caso novo): adiciona linha ao final da aba PRAZOS.

    As colunas D (PRAZO 1) e E (PRAZO 2) ficam vazias para preenchimento
    manual posterior, como especificado. Retorna o número da linha criada.
    """

    ws = wb[ABA_PRAZOS]
    linha = _proxima_linha_livre(ws, COL_PRAZOS_PROCESSO)

    ws.cell(row=linha, column=COL_PRAZOS_TRIBUNAL, value=caso.tribunal)
    ws.cell(row=linha, column=COL_PRAZOS_PROCESSO, value=caso.numero_processo)
    ws.cell(row=linha, column=COL_PRAZOS_CLIENTE, value=caso.cliente_x_parte)
    # D e E ficam em branco (preenchimento manual posterior).
    ws.cell(
        row=linha,
        column=COL_PRAZOS_INTERNO,
        value=f'=IF($D{linha}="","",WORKDAY($D{linha},-2))',
    )
    ws.cell(row=linha, column=COL_PRAZOS_ADVOGADO, value=caso.advogado)
    ws.cell(row=linha, column=COL_PRAZOS_STATUS, value="PENDENTE")
    observacoes = caso.observacoes
    if caso.segredo_justica:
        observacoes = "Segredo de Justiça" if not observacoes else f"{observacoes}; Segredo de Justiça"
    if observacoes:
        ws.cell(row=linha, column=COL_PRAZOS_OBSERVACOES, value=observacoes)
    return linha


@dataclass
class NovaPericia:
    cliente: str
    numero_processo: str
    pericia: str
    perito: str
    data: date
    horario: str
    local: str | None = None


def adicionar_linha_pericia(wb: Workbook, item: NovaPericia) -> int:
    """Seção 4: adiciona entrada na aba PERÍCIA (mesmo para casos recorrentes)."""

    ws = wb[ABA_PERICIA]
    linha = _proxima_linha_livre(ws, 2)
    ws.cell(row=linha, column=1, value=item.cliente)
    ws.cell(row=linha, column=2, value=item.numero_processo)
    ws.cell(row=linha, column=3, value=item.pericia)
    ws.cell(row=linha, column=4, value=item.perito)
    ws.cell(row=linha, column=5, value=item.data)
    ws.cell(row=linha, column=6, value=item.horario)
    if item.local:
        ws.cell(row=linha, column=7, value=item.local)
    return linha


@dataclass
class NovaAudiencia:
    cliente: str
    numero_processo: str
    evento: str
    area: str
    data: date
    horario: str


def processos_na_aba_audiencia(ws: Worksheet) -> set[str]:
    encontrados: set[str] = set()
    for linha in range(2, ws.max_row + 1):
        valor = ws.cell(row=linha, column=2).value
        if valor:
            encontrados.add(normalizar_processo(str(valor)))
    return encontrados


def adicionar_linha_audiencia(wb: Workbook, item: NovaAudiencia) -> int:
    """Seção 9: adiciona audiência futura ainda não cadastrada.

    Cliente é gravado em maiúsculas. AGENDADO(A) nunca é marcado "SIM"
    automaticamente (só manualmente, após lançar na agenda do Google).
    """

    ws = wb[ABA_AUDIENCIA]
    linha = _proxima_linha_livre(ws, 2)
    ws.cell(row=linha, column=1, value=item.cliente.upper())
    ws.cell(row=linha, column=2, value=item.numero_processo)
    ws.cell(row=linha, column=3, value=item.evento)
    ws.cell(row=linha, column=4, value=item.area)
    ws.cell(row=linha, column=5, value=item.data)
    ws.cell(row=linha, column=6, value=item.horario)
    # LINK, CONTATO DO CLIENTE, CIENTE, AGENDADO(A), CONFIRMAÇÃO,
    # TESTEMUNHAS, OBSERVAÇÕES ficam em branco (colunas 7 a 13).
    return linha


def salvar_com_seguranca(wb: Workbook, caminho_original: Path) -> None:
    """Seção 3: gravação segura do xlsx.

    Salva em arquivo temporário no mesmo diretório e substitui o original de
    forma atômica com ``os.replace``. Nunca chamar ``wb.save(caminho_original)``
    diretamente (bug observado de gravação que pode não persistir).
    """

    caminho_original = Path(caminho_original)
    fd, tmp_nome = tempfile.mkstemp(
        prefix=".tmp_write_", suffix=".xlsx", dir=caminho_original.parent
    )
    os.close(fd)
    tmp_caminho = Path(tmp_nome)
    try:
        wb.save(tmp_caminho)
        os.replace(tmp_caminho, caminho_original)
    finally:
        if tmp_caminho.exists():
            tmp_caminho.unlink(missing_ok=True)


def fazer_backup(caminho_original: Path, pasta_outputs: Path) -> Path:
    """Copia o xlsx original para a pasta de outputs antes de qualquer alteração."""

    pasta_outputs.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = pasta_outputs / f"PRAZOS_BECKER_2026_backup_{timestamp}.xlsx"
    shutil.copy2(caminho_original, destino)
    return destino


def conferir_celula(caminho: Path, aba: str, linha: int, coluna: int, valor_esperado) -> bool:
    """Reabre o arquivo do zero do disco e confere que a alteração persistiu."""

    wb = load_workbook(caminho, data_only=False)
    try:
        valor_real = wb[aba].cell(row=linha, column=coluna).value
        return valor_real == valor_esperado
    finally:
        wb.close()
