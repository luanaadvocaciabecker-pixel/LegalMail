import datetime as dt
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from legalmail_prazos.holidays import Calendario
from legalmail_prazos.legalmail_client import (
    AudienciaLegalmail,
    ItemEntrada,
    NovaTarefaLegalmail,
    TipoTarefaHistorico,
)
from legalmail_prazos.planilha import ABA_AUDIENCIA, ABA_PRAZOS
from legalmail_prazos.prazos import Prazo, RegimeContagem, calcular_prazo
from legalmail_prazos.rotina import (
    DecisaoItemEntrada,
    escolher_tipo_tarefa,
    escrever_audiencias_da_entrada,
    processar_parte1,
    processar_parte2,
    sugerir_pericia,
    triar_entrada,
)


@dataclass
class FakeLegalmailClient:
    historico_por_processo: dict[str, list[TipoTarefaHistorico]] = field(default_factory=dict)
    usuario_id_por_nome: dict[str, int] = field(default_factory=dict)
    tarefas_criadas: list[NovaTarefaLegalmail] = field(default_factory=list)
    processos_arquivados: list[str] = field(default_factory=list)
    processos_encarregados: list[tuple[str, int]] = field(default_factory=list)

    def listar_entrada(self):  # pragma: no cover - não usado nestes testes
        return []

    def historico_tarefas_do_processo(self, id_legalmail_processo: str):
        return self.historico_por_processo.get(id_legalmail_processo, [])

    def criar_tarefa(self, tarefa: NovaTarefaLegalmail) -> str:
        self.tarefas_criadas.append(tarefa)
        return f"tarefa-{len(self.tarefas_criadas)}"

    def arquivar_para_acervo(self, id_legalmail_processo: str) -> None:
        self.processos_arquivados.append(id_legalmail_processo)

    def localizar_usuario_por_nome(self, nome: str) -> int | None:
        return self.usuario_id_por_nome.get(nome)

    def encarregar_advogado(self, id_legalmail_processo: str, id_usuario: int) -> None:
        self.processos_encarregados.append((id_legalmail_processo, id_usuario))

    def listar_audiencias(self, *, a_partir_de: date):  # pragma: no cover
        return []


def _prazo_teste() -> Prazo:
    cal = Calendario(anos=(2026,), incluir_carnaval=False)
    return calcular_prazo(
        data_disponibilizacao=date(2026, 8, 24),
        quantidade_dias=10,
        regime=RegimeContagem.CIVEL_DIAS_UTEIS,
        calendario=cal,
    )


def test_escolher_tipo_tarefa_reaproveita_por_similaridade():
    historico = [
        TipoTarefaHistorico(tipo="CIENCIA", descricao_resumo="Ciência de despacho simples"),
        TipoTarefaHistorico(
            tipo="MANIFESTAÇÃO ENDEREÇO", descricao_resumo="Manifestar sobre endereço do réu"
        ),
    ]
    tipo = escolher_tipo_tarefa(historico, "Manifestar sobre o novo endereço informado")
    assert tipo == "MANIFESTAÇÃO ENDEREÇO"


def test_escolher_tipo_tarefa_sem_historico_retorna_none():
    assert escolher_tipo_tarefa([], "qualquer descrição") is None


def test_processar_parte1_caso_novo_cria_linha_tarefa_e_arquiva(planilha_sintetica: Path, tmp_path: Path):
    client = FakeLegalmailClient(
        historico_por_processo={
            "item-1": [
                TipoTarefaHistorico(tipo="CIENCIA", descricao_resumo="Ciência de decisão")
            ]
        }
    )
    item = ItemEntrada(
        id_legalmail="item-1",
        numero_processo="0001000-00.2026.5.12.0001",
        tribunal="TRT12 1G",
        cliente_x_parte="CLIENTE TESTE DOIS",
        conteudo_intimacao="Decisão para ciência.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    decisao = DecisaoItemEntrada(
        item=item,
        descricao_tarefa="Ciência de decisão proferida nos autos",
        prazo=_prazo_teste(),
        encarregado_override="BELTRANA SILVA",
    )

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    assert len(relatorio.casos_novos) == 1
    assert relatorio.casos_novos[0].numero_processo == "0001000-00.2026.5.12.0001"
    assert relatorio.tarefas_criadas == 1
    assert client.tarefas_criadas[0].tipo == "CIENCIA"
    assert client.processos_arquivados == ["item-1"]

    wb = load_workbook(planilha_sintetica)
    processos = {ws_row[1] for ws_row in wb[ABA_PRAZOS].iter_rows(min_row=9, values_only=True)}
    assert "0001000-00.2026.5.12.0001" in processos


def test_processar_parte1_encarrega_advogado_quando_usuario_e_encontrado(
    planilha_sintetica: Path, tmp_path: Path
):
    client = FakeLegalmailClient(usuario_id_por_nome={"BELTRANA SILVA": 42})
    item = ItemEntrada(
        id_legalmail="item-5",
        numero_processo="0009000-00.2026.5.12.0001",
        tribunal="TRT12 1G",
        cliente_x_parte="CLIENTE TESTE NOVE",
        conteudo_intimacao="Decisão para ciência.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    decisao = DecisaoItemEntrada(
        item=item,
        descricao_tarefa="Ciência de decisão proferida nos autos",
        prazo=_prazo_teste(),
        encarregado_override="BELTRANA SILVA",
    )

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.processos_encarregados == 1
    assert client.processos_encarregados == [("item-5", 42)]
    assert not any("encarregar" in limite for limite in relatorio.limitacoes)


def test_processar_parte1_reporta_limitacao_quando_usuario_nao_encontrado(
    planilha_sintetica: Path, tmp_path: Path
):
    client = FakeLegalmailClient()  # usuario_id_por_nome vazio
    item = ItemEntrada(
        id_legalmail="item-6",
        numero_processo="0010000-00.2026.5.12.0001",
        tribunal="TRT12 1G",
        cliente_x_parte="CLIENTE TESTE DEZ",
        conteudo_intimacao="Decisão para ciência.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    decisao = DecisaoItemEntrada(
        item=item,
        descricao_tarefa="Ciência de decisão proferida nos autos",
        prazo=_prazo_teste(),
        encarregado_override="ADVOGADO SEM CADASTRO NO LEGALMAIL",
    )

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.processos_encarregados == 0
    assert client.processos_encarregados == []
    assert any("encarregar" in limite for limite in relatorio.limitacoes)


class FakeLegalmailApiOnlyClient(FakeLegalmailClient):
    """Simula o adaptador real da API pública: sem tarefa, sem histórico."""

    def historico_tarefas_do_processo(self, id_legalmail_processo: str):
        raise NotImplementedError("sem histórico de tarefas na API pública")

    def criar_tarefa(self, tarefa: NovaTarefaLegalmail) -> str:
        raise NotImplementedError("sem endpoint de criar tarefa na API pública")


def test_processar_parte1_degrada_sem_crashar_quando_api_nao_suporta_tarefa(
    planilha_sintetica: Path, tmp_path: Path
):
    client = FakeLegalmailApiOnlyClient(usuario_id_por_nome={"BELTRANA SILVA": 42})
    item = ItemEntrada(
        id_legalmail="item-7",
        numero_processo="0011000-00.2026.5.12.0001",
        tribunal="TRT12 1G",
        cliente_x_parte="CLIENTE TESTE ONZE",
        conteudo_intimacao="Decisão para ciência.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    decisao = DecisaoItemEntrada(
        item=item,
        descricao_tarefa="Ciência de decisão proferida nos autos",
        prazo=_prazo_teste(),
        encarregado_override="BELTRANA SILVA",
    )

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    # A tarefa não pôde ser criada, mas o encaminhamento ao responsável e o
    # arquivamento continuam funcionando via API.
    assert relatorio.tarefas_criadas == 0
    assert relatorio.processos_encarregados == 1
    assert client.processos_arquivados == ["item-7"]
    assert any("tarefa" in limite for limite in relatorio.limitacoes)


def test_processar_parte1_caso_recorrente_nao_duplica_linha(planilha_sintetica: Path, tmp_path: Path):
    client = FakeLegalmailClient()
    item = ItemEntrada(
        id_legalmail="item-2",
        numero_processo="5000000-00.2025.8.24.0038",  # já existe na planilha sintética
        tribunal="TJSC 1G",
        cliente_x_parte="CLIENTE TESTE UM",
        conteudo_intimacao="Nova intimação no processo já conhecido.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    decisao = DecisaoItemEntrada(
        item=item,
        descricao_tarefa="Nova manifestação necessária",
        prazo=_prazo_teste(),
        encarregado_override="FULANA DE TAL",
    )

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.casos_novos == []
    assert relatorio.tarefas_criadas == 1

    wb = load_workbook(planilha_sintetica)
    processos = [
        row[1] for row in wb[ABA_PRAZOS].iter_rows(min_row=9, values_only=True) if row[1]
    ]
    assert processos.count("5000000-00.2025.8.24.0038") == 1


def test_processar_parte1_conteudo_bloqueado_nao_inventa_dados(planilha_sintetica: Path, tmp_path: Path):
    client = FakeLegalmailClient()
    item = ItemEntrada(
        id_legalmail="item-3",
        numero_processo="0005000-00.2026.8.24.0038",
        tribunal="TJSC 1G",
        cliente_x_parte="CLIENTE TESTE CINCO",
        conteudo_intimacao="",
        data_disponibilizacao=date(2026, 8, 24),
        segredo_justica=True,
        conteudo_acessivel=False,
    )
    decisao = DecisaoItemEntrada(item=item, descricao_tarefa="", prazo=_prazo_teste())

    relatorio = processar_parte1(
        decisoes=[decisao],
        client=client,
        caminho_planilha=planilha_sintetica,
        caminho_prazos_nums_json=tmp_path / "prazos_nums.json",
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.tarefas_criadas == 0
    assert client.tarefas_criadas == []
    assert relatorio.limitacoes  # a limitação foi reportada


def test_relatorio_texto_sem_dois_pontos():
    from legalmail_prazos.rotina import RelatorioConciliacao, ResultadoCasoNovo

    relatorio = RelatorioConciliacao(
        casos_novos=[
            ResultadoCasoNovo(numero_processo="0006000-00.2026.8.24.0038", cliente="CLIENTE TESTE SEIS", tribunal="TJSC 1G")
        ],
        tarefas_criadas=1,
    )
    assert ":" not in relatorio.texto()


def test_processar_parte2_adiciona_audiencia_nova(planilha_sintetica: Path, tmp_path: Path):
    audiencia = AudienciaLegalmail(
        id_legalmail="aud-1",
        numero_processo="0007000-00.2026.5.12.0030",
        cliente="cliente teste sete",
        evento="AUDIÊNCIA DE CONCILIAÇÃO",
        area="TRABALHISTA",
        data=date(2026, 10, 10),
        horario=dt.time(14, 0),
    )
    relatorio = processar_parte2(
        audiencias_legalmail=[audiencia],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )
    assert relatorio.audiencias_adicionadas == 1

    wb = load_workbook(planilha_sintetica)
    ws = wb[ABA_AUDIENCIA]
    linha = ws.max_row
    assert ws.cell(row=linha, column=1).value == "CLIENTE TESTE SETE"
    assert ws.cell(row=linha, column=10).value is None


def test_processar_parte2_nao_duplica_audiencia_existente(planilha_sintetica: Path, tmp_path: Path):
    audiencia = AudienciaLegalmail(
        id_legalmail="aud-2",
        numero_processo="0008000-00.2026.5.12.0030",
        cliente="cliente teste oito",
        evento="AUDIÊNCIA DE INSTRUÇÃO",
        area="TRABALHISTA",
        data=date(2026, 10, 12),
        horario=dt.time(9, 0),
    )
    processar_parte2(
        audiencias_legalmail=[audiencia],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )
    relatorio2 = processar_parte2(
        audiencias_legalmail=[audiencia],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )
    assert relatorio2.audiencias_adicionadas == 0


def _item_entrada(**overrides) -> ItemEntrada:
    base = dict(
        id_legalmail="item-x",
        numero_processo="0099000-00.2026.5.12.0030",
        tribunal="TRT12 1G",
        cliente_x_parte="CLIENTE TESTE TRIAGEM",
        conteudo_intimacao="Manifeste-se em 15 dias sobre a contestação.",
        data_disponibilizacao=date(2026, 8, 24),
    )
    base.update(overrides)
    return ItemEntrada(**base)


def test_triar_entrada_separa_audiencia_dos_demais():
    item_audiencia = _item_entrada(id_legalmail="a1", tipo="Audiência")
    item_prazo = _item_entrada(id_legalmail="a2", tipo="Intimação")

    audiencias, demais = triar_entrada([item_audiencia, item_prazo])

    assert audiencias == [item_audiencia]
    assert demais == [item_prazo]


def test_sugerir_pericia_retorna_none_sem_mencao():
    item = _item_entrada(conteudo_intimacao="Manifeste-se em 15 dias.")
    assert sugerir_pericia(item) is None


def test_sugerir_pericia_extrai_data_quando_disponivel():
    item = _item_entrada(
        conteudo_intimacao="Nomeio perito e designo perícia para o dia 10/11/2026 às 09:00."
    )
    pericia = sugerir_pericia(item)
    assert pericia is not None
    assert pericia.numero_processo == item.numero_processo
    assert pericia.data == date(2026, 11, 10)
    assert pericia.horario == "09:00"


def test_escrever_audiencias_da_entrada_extrai_e_grava(planilha_sintetica: Path, tmp_path: Path):
    item = _item_entrada(
        tipo="Audiência",
        conteudo_intimacao="Designo audiência de instrução para o dia 06/05/2026 às 14:30.",
    )

    relatorio = escrever_audiencias_da_entrada(
        itens_audiencia=[item],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.audiencias_adicionadas == 1
    assert not relatorio.limitacoes

    wb = load_workbook(planilha_sintetica)
    ws = wb[ABA_AUDIENCIA]
    linha = ws.max_row
    assert ws.cell(row=linha, column=1).value == "CLIENTE TESTE TRIAGEM"
    assert ws.cell(row=linha, column=3).value == "AUDIÊNCIA DE INSTRUÇÃO"
    assert ws.cell(row=linha, column=4).value == "TRABALHISTA"
    assert ws.cell(row=linha, column=6).value == "14:30"


def test_escrever_audiencias_da_entrada_sinaliza_quando_nao_extrai_data(
    planilha_sintetica: Path, tmp_path: Path
):
    item = _item_entrada(
        tipo="Audiência",
        numero_processo="0099001-00.2026.5.12.0030",
        conteudo_intimacao="Audiência designada, aguardar publicação da pauta.",
    )

    relatorio = escrever_audiencias_da_entrada(
        itens_audiencia=[item],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio.audiencias_adicionadas == 1
    assert relatorio.limitacoes  # sinalizado para confirmação manual

    wb = load_workbook(planilha_sintetica)
    ws = wb[ABA_AUDIENCIA]
    linha = ws.max_row
    assert ws.cell(row=linha, column=5).value is None  # DATA em branco, não inventada


def test_escrever_audiencias_da_entrada_nao_duplica(planilha_sintetica: Path, tmp_path: Path):
    item = _item_entrada(
        tipo="Audiência",
        numero_processo="0099002-00.2026.5.12.0030",
        conteudo_intimacao="Audiência de conciliação designada para 12/12/2026 às 10:00.",
    )

    escrever_audiencias_da_entrada(
        itens_audiencia=[item],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )
    relatorio2 = escrever_audiencias_da_entrada(
        itens_audiencia=[item],
        caminho_planilha=planilha_sintetica,
        pasta_outputs=tmp_path / "outputs",
    )

    assert relatorio2.audiencias_adicionadas == 0
