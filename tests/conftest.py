from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from legalmail_prazos.planilha import (
    ABA_ATIVOS_ATUAL,
    ABA_AUDIENCIA,
    ABA_PERICIA,
    ABA_PRAZOS,
)


@pytest.fixture
def planilha_sintetica(tmp_path: Path) -> Path:
    """Cria uma planilha .xlsx sintética com a mesma estrutura da real, sem
    nenhum dado de cliente verdadeiro, para uso isolado nos testes."""

    wb = Workbook()
    wb.remove(wb.active)

    ws_prazos = wb.create_sheet(ABA_PRAZOS)
    ws_prazos.append(["RESUMO GERAL DE PRAZOS"])
    ws_prazos.append(["Data de referência: 01/01/2026"])
    ws_prazos.append([None, "LEGENDA"])
    ws_prazos.append([None, "PRAZO 1", "ULTIMO DIA", None, "FULANA DE TAL"])
    ws_prazos.append([None, "PRAZO 2", "PRAZO ANTECIPADO", None, "BELTRANA SILVA"])
    ws_prazos.append([None, "INTERNO", "2 DIAS ANTES", None, "CICRANA SOUZA"])
    ws_prazos.append([None])
    ws_prazos.append(
        ["TRIBUNAL", "Nº DO PROCESSO", "CLIENTE", "PRAZO 1", "PRAZO 2", "INTERNO", "ADVOGADO", "STATUS", "OBSERVAÇÕES"]
    )
    ws_prazos.append(
        [
            "TJSC 1G",
            "5000000-00.2025.8.24.0038",
            "CLIENTE TESTE UM",
            datetime(2026, 6, 1),
            "=IF($D9=\"\",\"\",WORKDAY($D9,-10))",
            "=IF($D9=\"\",\"\",WORKDAY($D9,-2))",
            "FULANA DE TAL",
            "PROTOCOLADO",
            None,
        ]
    )

    ws_ativos = wb.create_sheet(ABA_ATIVOS_ATUAL)
    ws_ativos.append(
        [
            "Nº PROCESSO",
            "DATA AJUIZAMENTO",
            "DATA ÚLTIMO MOVIMENTO",
            "TRIBUNAL",
            "GRAU",
            "CLASSE",
            "ASSUNTO PRINCIPAL",
            "PARTES",
            "ÓRGÃO JULGADOR",
            "PASTA (PROJURIS)",
            "ADVOGADA ATUAL",
        ]
    )
    ws_ativos.append(
        ["5000000-00.2025.8.24.0038", None, None, "TJSC", "1º Grau", None, None, None, None, 1, "Fulana"]
    )
    ws_ativos.append(
        ["5000001-11.2026.8.24.0038", None, None, "TJSC", "1º Grau", None, None, None, None, 2, "Beltrana"]
    )

    wb.create_sheet(ABA_PERICIA).append(
        ["CLIENTE", "N° DO PROCESSO", "PERÍCIA", "PERITO", "DATA", "HORÁRIO", "LOCAL"]
    )

    wb.create_sheet(ABA_AUDIENCIA).append(
        [
            "CLIENTE",
            "N° DO PROCESSO",
            "EVENTO",
            "ÁREA",
            "DATA",
            "HORÁRIO",
            "LINK",
            "CONTATO DO CLIENTE",
            "CIENTE",
            "AGENDADO(A)",
            "CONFIRMAÇÃO",
            "TESTEMUNHAS",
            "OBSERVAÇÕES",
        ]
    )

    caminho = tmp_path / "PRAZOS_TESTE.xlsx"
    wb.save(caminho)
    return caminho
