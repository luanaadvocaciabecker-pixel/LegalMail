from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from legalmail_prazos.planilha import (
    ABA_AUDIENCIA,
    ABA_PERICIA,
    ABA_PRAZOS,
    NovaAudiencia,
    NovaPericia,
    NovoCasoPrazos,
    adicionar_linha_audiencia,
    adicionar_linha_pericia,
    adicionar_linha_prazos,
    carregar_processos_conhecidos,
    conferir_celula,
    eh_caso_novo,
    fazer_backup,
    localizar_advogado_por_processo,
    normalizar_processo,
    processos_na_aba_audiencia,
    processos_na_aba_prazos,
    salvar_com_seguranca,
    salvar_processos_conhecidos,
)


def test_normalizar_processo_remove_pontuacao_e_nbsp():
    assert normalizar_processo("5000000-00.2025.8.24.0038") == "50000000020258240038"
    assert normalizar_processo("\xa0\xa00000609-24.2026.5.12.0030") == "00006092420265120030"


def test_processos_na_aba_prazos(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    processos = processos_na_aba_prazos(wb[ABA_PRAZOS])
    assert "50000000020258240038" in processos


def test_eh_caso_novo():
    conhecidos = {normalizar_processo("5000000-00.2025.8.24.0038")}
    assert eh_caso_novo("5000000-00.2025.8.24.0038", conhecidos) is False
    assert eh_caso_novo("5099999-99.2026.8.24.0038", conhecidos) is True


def test_localizar_advogado_por_processo_sem_mapa(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    advogado = localizar_advogado_por_processo(wb, "5000001-11.2026.8.24.0038")
    assert advogado == "Beltrana"


def test_localizar_advogado_por_processo_com_mapa(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    mapa = {"Beltrana": "BELTRANA SILVA"}
    advogado = localizar_advogado_por_processo(
        wb, "5000001-11.2026.8.24.0038", mapa_abreviacao_para_nome_completo=mapa
    )
    assert advogado == "BELTRANA SILVA"


def test_localizar_advogado_processo_desconhecido_retorna_none(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    assert localizar_advogado_por_processo(wb, "9999999-99.2099.8.24.0038") is None


def test_adicionar_linha_prazos_novo_caso(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    linha = adicionar_linha_prazos(
        wb,
        NovoCasoPrazos(
            tribunal="TRT12 1G",
            numero_processo="0001000-00.2026.5.12.0001",
            cliente_x_parte="CLIENTE TESTE DOIS",
            advogado="BELTRANA SILVA",
            segredo_justica=True,
        ),
    )
    ws = wb[ABA_PRAZOS]
    assert ws.cell(row=linha, column=1).value == "TRT12 1G"
    assert ws.cell(row=linha, column=2).value == "0001000-00.2026.5.12.0001"
    assert ws.cell(row=linha, column=4).value is None  # PRAZO 1 fica em branco
    assert ws.cell(row=linha, column=6).value == f'=IF($D{linha}="","",WORKDAY($D{linha},-2))'
    assert ws.cell(row=linha, column=7).value == "BELTRANA SILVA"
    assert ws.cell(row=linha, column=8).value == "PENDENTE"
    assert ws.cell(row=linha, column=9).value == "Segredo de Justiça"


def test_adicionar_linha_pericia(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    linha = adicionar_linha_pericia(
        wb,
        NovaPericia(
            cliente="CLIENTE TESTE DOIS",
            numero_processo="0001000-00.2026.5.12.0001",
            pericia="MÉDICA",
            perito="PERITO TESTE",
            data=date(2026, 10, 1),
            horario="09:00",
        ),
    )
    ws = wb[ABA_PERICIA]
    assert ws.cell(row=linha, column=1).value == "CLIENTE TESTE DOIS"
    assert ws.cell(row=linha, column=3).value == "MÉDICA"


def test_adicionar_linha_audiencia_maiusculo_e_nao_marca_agendado(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    linha = adicionar_linha_audiencia(
        wb,
        NovaAudiencia(
            cliente="cliente teste tres",
            numero_processo="0002000-00.2026.5.12.0001",
            evento="AUDIÊNCIA DE INSTRUÇÃO",
            area="TRABALHISTA",
            data=date(2026, 10, 5),
            horario="14:00",
        ),
    )
    ws = wb[ABA_AUDIENCIA]
    assert ws.cell(row=linha, column=1).value == "CLIENTE TESTE TRES"
    assert ws.cell(row=linha, column=10).value is None  # AGENDADO(A) nunca marcado automaticamente


def test_processos_na_aba_audiencia_ignora_linha_vazia(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    assert processos_na_aba_audiencia(wb[ABA_AUDIENCIA]) == set()


def test_salvar_com_seguranca_e_conferir_celula(planilha_sintetica: Path):
    wb = load_workbook(planilha_sintetica)
    linha = adicionar_linha_prazos(
        wb,
        NovoCasoPrazos(
            tribunal="TJSC 1G",
            numero_processo="0003000-00.2026.8.24.0038",
            cliente_x_parte="CLIENTE TESTE QUATRO",
            advogado="FULANA DE TAL",
        ),
    )
    salvar_com_seguranca(wb, planilha_sintetica)

    assert conferir_celula(
        planilha_sintetica, ABA_PRAZOS, linha, 2, "0003000-00.2026.8.24.0038"
    )
    assert not conferir_celula(planilha_sintetica, ABA_PRAZOS, linha, 2, "número errado")

    wb_reaberto = load_workbook(planilha_sintetica)
    assert wb_reaberto[ABA_PRAZOS].cell(row=linha, column=2).value == "0003000-00.2026.8.24.0038"


def test_fazer_backup_copia_arquivo(planilha_sintetica: Path, tmp_path: Path):
    outputs = tmp_path / "outputs"
    destino = fazer_backup(planilha_sintetica, outputs)
    assert destino.exists()
    assert destino.read_bytes() == planilha_sintetica.read_bytes()


def test_processos_conhecidos_roundtrip(tmp_path: Path):
    caminho = tmp_path / "prazos_nums.json"
    processos = {"12345678901234567890", "98765432109876543210"}
    salvar_processos_conhecidos(caminho, processos)
    assert carregar_processos_conhecidos(caminho) == processos
